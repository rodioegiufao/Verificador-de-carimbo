import PyPDF2
import os
import re
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from io import BytesIO
import pandas as pd

# Configuração da página Streamlit
st.set_page_config(page_title="Analisador de Carimbos PDF", page_icon="📄", layout="wide")

# Título da aplicação
st.title("📄 Analisador de Carimbos em PDFs")
st.markdown("Esta ferramenta verifica a presença de palavras-chave em arquivos PDF e gera um relatório.")

# Dados FIXOS dos engenheiros e CREAs (sempre serão pesquisados)
ENGENHEIROS_CREAS_FIXOS = {
    "RODRIGO DAMASCENO NASCIMENTO": ["0920192912", "092019291-2"],
    "JÂNIO RIBEIRO LOPES": ["0912111810", "091211181-0"],
    "FLAVIO SORDI": ["2201136580"],
    "RITHELLY LOBATO": ["A278773-3", "A2787733"],
    "SALOMÃO": ["0401863549", "040186354-9"]
}

# Mapeamento de códigos de projeto para descrições
MAPEAMENTO_PROJETOS = {
    "ECX": "PROJETO ELÉTRICO DE BAIXA",
    "ILUX": "PROJETO DE ILUMINAÇÃO EXTERNA",
    "CFTV": "PROJETO DE CFTV",
    "CAB": "PROJETO DE CABEAMENTO",
    "SOM": "PROJETO DE SONORIZAÇÃO",
    "SUB": "PROJETO DE SUBESTAÇÃO",
    "SPDA": "PROJETO DE SPDA",
    "TEF": "PROJETO DE TELEFONIA",
    "ALI": "PROJETO ELÉTRICO DE BAIXA",
    "TUG": "PROJETO ELÉTRICO DE BAIXA",
    "ILU": "PROJETO ELÉTRICO DE BAIXA",
    "EME": "PROJETO ELÉTRICO DE BAIXA",
    "FOT": "PROJETO ELÉTRICO FOTOVOLTAICO",
    "LEV": "LEVANTAMENTO TOPOGRÁFICO",
    "EST": "ESTRUTURA DE CONCRETO ARMADO",
    "FUN": "ESTRUTURA DE CONCRETO ARMADO", 
    "EMT": "ESTRUTURA METÁLICA",
    "DRE": "PROJETO DE DRENAGEM",
    "PAV": "PROJETO DE PAVIMENTAÇÃO",
    "REG": "PROJETO DE REDE DE ESGOTO",
    "TER": "PROJETO DE TERRAPLENAGEM",
    "CANT": "PROJETO DE CANTEIRO DE OBRAS",
    "HID": "PROJETO DE INSTALAÇÕES HIDRÁULICAS",
    "IRRI": "PROJETO DE IRRIGAÇÃO",
    "SAN": "PROJETO DE INSTALAÇÕES SANITÁRIAS",
    "PLU": "PROJETO DE SISTEMA DE REDES DE ÁGUAS",
    "INC": "PROJETO DE PREVENÇÃO E COMBATE A INCÊNDIO",
    "GLP": "PROJETO DE INSTALAÇÕES DE GASES GLP",
    "CLI": "PROJETO DE INSTALAÇÕES DE GASES GLP",
    "EXA": "PROJETO DE EXAUSTÃO"
}

# Criar lista FIXA de palavras-chave dos engenheiros (sempre serão pesquisadas)
PALAVRAS_CHAVE_ENGENHEIROS = []
for engenheiro, creas in ENGENHEIROS_CREAS_FIXOS.items():
    PALAVRAS_CHAVE_ENGENHEIROS.append(engenheiro)
    PALAVRAS_CHAVE_ENGENHEIROS.extend(creas)

# Palavras-chave padrão adicionais (projeto específico)
PALAVRAS_CHAVE_PADRAO = [
    "IPER",
    "CONSTRUÇÃO DA SEDE DO INSTITUTO DE PREVIDÊNCIA DO ESTADO",
    "DE RORAIMA - IPER",
    "AGOSTO",
    "2025",
    "RUA",
    "CC-22",
    "LOTE: 712 - REM.", 
    "LAURA MOREIRA",
    "69318-105",
    "BOA VISTA",
    "RR",
    "2.220,32",
    "2.654,11",
    "SAUDE",
    "SAÚDE"
]

# Criar abas para diferentes funcionalidades
tab1, tab2 = st.tabs(["Analisador de PDF", "Tutorial em Vídeo"])

with tab1:
    # Sidebar para configurações
    with st.sidebar:
        st.header("Configurações")
        
        # Upload de arquivos PDF
        uploaded_files = st.file_uploader("Selecione os arquivos PDF", type="pdf", accept_multiple_files=True)
        
        # Entrada de palavras-chave
        st.subheader("Palavras-chave")
        
        # Informar que engenheiros serão sempre pesquisados
        st.info("🔍 Engenheiros e CREAs serão SEMPRE pesquisados:")
        with st.expander("Ver engenheiros e CREAs fixos"):
            for engenheiro, creas in ENGENHEIROS_CREAS_FIXOS.items():
                st.write(f"**{engenheiro}**: {', '.join(creas)}")
        
        # Mostrar mapeamento de projetos
        st.info("🗂️ Mapeamento de códigos de projeto:")
        with st.expander("Ver mapeamento de projetos"):
            for codigo, descricao in MAPEAMENTO_PROJETOS.items():
                st.write(f"**{codigo}**: {descricao}")
        
        # Campo para palavras-chave adicionais (projeto específico)
        keywords_input = st.text_area(
            "Insira palavras-chave adicionais do projeto (uma por linha)", 
            value="\n".join(PALAVRAS_CHAVE_PADRAO),
            height=200,
            help="Estas palavras-chave serão pesquisadas além dos engenheiros e CREAs"
        )
        
        # Opções adicionais
        st.subheader("Opções")
        check_filename = st.checkbox("Verificar nome do arquivo no conteúdo", value=True)
        check_sheet_number = st.checkbox("Verificar número da prancha no conteúdo", value=True)
        check_projeto = st.checkbox("Verificar descrição do projeto no conteúdo", value=True,
                                   help="Verifica se a descrição do projeto está presente no PDF")
        
        # Botão para iniciar análise
        analyze_button = st.button("Iniciar Análise", type="primary")

    # Função para extrair o número da prancha do nome do arquivo
    def extrair_numero_prancha(nome_arquivo):
        # Remove a extensão e possíveis sufixos como "_assinado"
        nome_sem_ext = os.path.splitext(nome_arquivo)[0].replace("_assinado", "")
        
        # Procura por padrões como _01_07 ou -01-07 no nome
        padroes = [
            r'[_\-](\d{2})[_\-](\d{2})(?:\..*)?$',
            r'[_\-](\d{2})[_\-](\d{3})(?:\..*)?$',
            r'[_\-](\d{3})[_\-](\d{3})(?:\..*)?$'
        ]
        
        for padrao in padroes:
            correspondencia = re.search(padrao, nome_sem_ext)
            if correspondencia:
                return f"{correspondencia.group(1)} {correspondencia.group(2)}"
        
        return None

    # Função para verificar se o arquivo está assinado pelo nome
    def verificar_assinatura_nome(nome_arquivo):
        # Verifica se o nome do arquivo contém "assinado" (case insensitive)
        return "assinado" in nome_arquivo.lower()

    # Função para extrair o código do projeto do nome do arquivo - CORRIGIDA
    def extrair_codigo_projeto(nome_arquivo):
        # Padrão mais flexível: PRJ-XXX- (onde XXX é o código do projeto)
        padrao = r'PRJ-([A-Z]+)-'
        correspondencia = re.search(padrao, nome_arquivo)
        if correspondencia:
            return correspondencia.group(1)
        return None

    # Processamento quando o botão é clicado
    if analyze_button and uploaded_files:
        # Preparar palavras-chave (engenheiros FIXOS + palavras adicionais do usuário)
        palavras_chave_adicionais = [linha.strip() for linha in keywords_input.split('\n') if linha.strip()]
        
        # Combinar palavras-chave fixas dos engenheiros com as adicionais
        todas_palavras_chave = PALAVRAS_CHAVE_ENGENHEIROS + palavras_chave_adicionais
        
        # Dicionário para armazenar os dados de cada PDF
        resultados = {}
        
        # Barra de progresso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Processar cada arquivo PDF
        for i, uploaded_file in enumerate(uploaded_files):
            # Atualizar barra de progresso
            progress = (i + 1) / len(uploaded_files)
            progress_bar.progress(progress)
            status_text.text(f"Processando {i+1} de {len(uploaded_files)}: {uploaded_file.name}")
            
            # Extrair o nome do arquivo sem a extensão
            nome_arquivo = os.path.splitext(uploaded_file.name)[0]
            
            # Verificar se o arquivo está assinado pelo nome
            assinado_pelo_nome = verificar_assinatura_nome(uploaded_file.name)
            
            # Extrair o número da prancha (removendo "_assinado" se existir)
            nome_sem_assinado = nome_arquivo.replace("_assinado", "")
            numero_prancha = extrair_numero_prancha(nome_sem_assinado)
            
            # Extrair o código do projeto
            codigo_projeto = extrair_codigo_projeto(uploaded_file.name)
            descricao_projeto = MAPEAMENTO_PROJETOS.get(codigo_projeto, "Desconhecido") if codigo_projeto else "Não identificado"
            
            # Lista para armazenar os dados encontrados no PDF atual
            dados_carimbo = []
            nome_arquivo_encontrado = False
            prancha_encontrada = False
            projeto_encontrado = False
            
            # Ler o conteúdo del PDF
            try:
                leitor = PyPDF2.PdfReader(uploaded_file)
                
                for pagina in leitor.pages:
                    texto_extraido = pagina.extract_text()
                    
                    if texto_extraido:
                        texto_extraido = texto_extraido.replace("\n", " ")  # Remover quebras de linha
                        
                        # Verificar se o nome do arquivo está no texto da página
                        if check_filename and nome_sem_assinado in texto_extraido:
                            nome_arquivo_encontrado = True
                        
                        # Verificar se o número da prancha está no texto
                        if check_sheet_number and numero_prancha:
                            if (numero_prancha.replace(" ", "_") in texto_extraido or 
                                numero_prancha.replace(" ", "-") in texto_extraido or 
                                numero_prancha in texto_extraido):
                                prancha_encontrada = True
                        
                        # Verificar se a descrição do projeto está no texto
                        if check_projeto and codigo_projeto and descricao_projeto != "Desconhecido":
                            if descricao_projeto in texto_extraido:
                                projeto_encontrado = True
                        
                        # Verificar palavras-chave FIXAS dos engenheiros
                        for palavra in PALAVRAS_CHAVE_ENGENHEIROS:
                            if palavra in texto_extraido and palavra not in dados_carimbo:
                                dados_carimbo.append(palavra)
                        
                        # Verificar palavras-chave adicionais do projeto
                        for palavra in palavras_chave_adicionais:
                            if palavra in texto_extraido and palavra not in dados_carimbo:
                                dados_carimbo.append(palavra)
            
            except Exception as e:
                st.error(f"Erro ao processar {uploaded_file.name}: {str(e)}")
                continue
            
            # Armazenar os dados no dicionário
            resultados[uploaded_file.name] = {
                'dados_carimbo': dados_carimbo,
                'nome_arquivo_encontrado': nome_arquivo_encontrado,
                'prancha_encontrada': prancha_encontrada,
                'assinado_pelo_nome': assinado_pelo_nome,
                'projeto_encontrado': projeto_encontrado,
                'codigo_projeto': codigo_projeto,
                'descricao_projeto': descricao_projeto,
                'numero_prancha': numero_prancha,
                'nome_arquivo': nome_arquivo
            }
        
        # Limpar barra de progresso
        progress_bar.empty()
        status_text.empty()
        
        # Criar DataFrame com os resultados
        dados_tabela = []
        for nome_arquivo, dados in resultados.items():
            dados_tabela.append({
                "Prancha": dados['numero_prancha'] if dados['numero_prancha'] else "Não identificado",
                "Código Projeto": dados['codigo_projeto'] if dados['codigo_projeto'] else "Não identificado",
                "Descrição Projeto": dados['descricao_projeto'],
                "Palavras-chave encontradas": ", ".join(dados['dados_carimbo']) if dados['dados_carimbo'] else "Nenhuma",
                "Nome do Arquivo": nome_arquivo,
                "Número da Prancha": dados['numero_prancha'] if dados['numero_prancha'] else "Não identificado",
                "Nome encontrado": "Sim" if dados['nome_arquivo_encontrado'] else "Não",
                "Prancha encontrada": "Sim" if dados['prancha_encontrada'] else "Não",
                "Arquivo assinado": "Sim" if dados['assinado_pelo_nome'] else "Não",
                "Projeto encontrado": "Sim" if dados['projeto_encontrado'] else "Não"
            })
        
        df = pd.DataFrame(dados_tabela)
        
        # Exibir resultados
        st.subheader("Resultados da Análise")
        st.dataframe(df)
        
        # Estatísticas
        st.subheader("Estatísticas")
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("Total de Arquivos", len(uploaded_files))
        with col2:
            st.metric("Arquivos com Nome Encontrado", 
                     sum(1 for dados in resultados.values() if dados['nome_arquivo_encontrado']))
        with col3:
            st.metric("Arquivos com Prancha Encontrada", 
                     sum(1 for dados in resultados.values() if dados['prancha_encontrada']))
        with col4:
            st.metric("Arquivos Assinados", 
                     sum(1 for dados in resultados.values() if dados['assinado_pelo_nome']))
        with col5:
            st.metric("Projetos Encontrados", 
                     sum(1 for dados in resultados.values() if dados['projeto_encontrado']))
        
        # Detalhamento dos engenheiros encontrados
        st.subheader("Engenheiros Encontrados")
        engenheiros_encontrados = {}
        for dados in resultados.values():
            for palavra in dados['dados_carimbo']:
                for engenheiro in ENGENHEIROS_CREAS_FIXOS:
                    if palavra in [engenheiro] + ENGENHEIROS_CREAS_FIXOS[engenheiro]:
                        engenheiros_encontrados[engenheiro] = engenheiros_encontrados.get(engenheiro, 0) + 1
        
        if engenheiros_encontrados:
            for engenheiro, count in engenheiros_encontrados.items():
                st.write(f"• **{engenheiro}**: encontrado em {count} arquivo(s)")
        else:
            st.write("Nenhum engenheiro encontrado nos arquivos analisados")
        
        # Detalhamento dos projetos
        st.subheader("Detalhamento dos Projetos")
        projetos_encontrados = {}
        for dados in resultados.values():
            codigo = dados['codigo_projeto'] or "Desconhecido"
            descricao = dados['descricao_projeto']
            if codigo != "Desconhecido":
                chave = f"{codigo} - {descricao}"
                projetos_encontrados[chave] = projetos_encontrados.get(chave, 0) + 1
        
        if projetos_encontrados:
            for projeto, count in projetos_encontrados.items():
                st.write(f"• **{projeto}**: {count} arquivo(s)")
        else:
            st.write("Nenhum projeto identificado nos arquivos analisados")
        
        # Criar planilha Excel para download
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados PDF"
        
        # Cabeçalhos (sem a coluna "Prancha" que é redundante)
        cabecalhos = ["Código Projeto", "Descrição Projeto", "Palavras-chave encontradas", 
                      "Nome do Arquivo", "Número da Prancha", "Nome encontrado", 
                      "Prancha encontrada", "Arquivo assinado", "Projeto encontrado"]
        
        # Escrever cabeçalhos em negrito
        for col, cabecalho in enumerate(cabecalhos, start=1):
            celula = ws.cell(row=1, column=col, value=cabecalho)
            celula.font = Font(bold=True)
        
        # Preencher os dados (pulando a coluna "Prancha" que é redundante)
        for row, dados in enumerate(dados_tabela, start=2):
            # Ignorar a primeira coluna ("Prancha") e escrever a partir da segunda
            ws.cell(row=row, column=1, value=dados["Código Projeto"])
            ws.cell(row=row, column=2, value=dados["Descrição Projeto"])
            ws.cell(row=row, column=3, value=dados["Palavras-chave encontradas"])
            ws.cell(row=row, column=4, value=dados["Nome do Arquivo"])
            ws.cell(row=row, column=5, value=dados["Número da Prancha"])
            ws.cell(row=row, column=6, value=dados["Nome encontrado"])
            ws.cell(row=row, column=7, value=dados["Prancha encontrada"])
            ws.cell(row=row, column=8, value=dados["Arquivo assinado"])
            ws.cell(row=row, column=9, value=dados["Projeto encontrado"])
        
        # Aplicar formatação condicional para as colunas de Sim/Não
        verde_claro = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        vermelho_claro = PatternFill(start_color="FF2C2B", end_color="FF2C2B", fill_type="solid")
        
        # Colunas para formatar (6: Nome encontrado, 7: Prancha encontrada, 8: Arquivo assinado, 9: Projeto encontrado)
        colunas_para_formatar = [6, 7, 8, 9]
        
        for coluna in colunas_para_formatar:
            for row in range(2, len(dados_tabela) + 2):
                celula = ws.cell(row=row, column=coluna)
                if celula.value == "Sim":
                    celula.fill = verde_claro
                elif celula.value == "Não":
                    celula.fill = vermelho_claro
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar para um buffer em memória
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Botão para download
        st.download_button(
            label="📥 Baixar Resultados em Excel",
            data=excel_buffer,
            file_name="resultados_analise.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.success("Análise concluída com sucesso! ✅")

    elif analyze_button and not uploaded_files:
        st.error("Por favor, selecione pelo menos um arquivo PDF.")

    # Instruções iniciais
    else:
        st.info("👈 Use a barra lateral para configurar e iniciar a análise.")
        st.markdown("""
        ### Como usar:
        1. Selecione os arquivos PDF que deseja analisar
        2. Insira palavras-chave adicionais do projeto (opcional)
        3. Ajuste as opções conforme necessário
        4. Clique em 'Iniciar Análise'
        
        ### ⚠️ Importante:
        - **Engenheiros e CREAs serão SEMPRE pesquisados** automaticamente
        - As palavras-chave adicionais são para complementar a busca
        - Os engenheiros pré-configurados não podem ser desativados
        - A verificação de assinatura é feita pelo nome do arquivo (presença de "assinado")
        - O código do projeto é extraído automaticamente del nome do arquivo
        - A descrição do projeto é mapeada automaticamente e verificada no conteúdo
        """)

with tab2:
    st.header("📹 Tutorial em Vídeo")
    st.markdown("""
    Assista ao vídeo abaixo para aprender como usar o analisador de PDF:
    """)
    
    # Inserir o vídeo do YouTube
    video_url = "https://youtu.be/GB6hvQPODCw"
    st.video(video_url)
    
    st.markdown("""
    ### Pontos abordados no tutorial:
    - Como preparar seus arquivos PDF para análise
    - Como usar palavras-chave adicionais do projeto
    - Interpretação dos resultados
    - Como usar o relatório em Excel gerado
    - Entendendo a busca automática por engenheiros e CREAs
    - Entendendo o mapeamento automático de códigos de projeto
    
    ### Engenheiros e CREAs configurados (busca automática):
    """)
    
    # Tabela com engenheiros e CREAs
    dados_engenheiros = []
    for engenheiro, creas in ENGENHEIROS_CREAS_FIXOS.items():
        dados_engenheiros.append({
            "Engenheiro": engenheiro,
            "CREA(s)": ", ".join(creas)
        })
    
    st.table(pd.DataFrame(dados_engenheiros))
    
    st.markdown("""
    ### Mapeamento de códigos de projeto:
    """)
    
    # Tabela com mapeamento de projetos
    dados_projetos = []
    for codigo, descricao in MAPEAMENTO_PROJETOS.items():
        dados_projetos.append({
            "Código": codigo,
            "Descrição": descricao
        })
    
    st.table(pd.DataFrame(dados_projetos))
    
    st.markdown("""
    ### ⚠️ Importante:
    Estes engenheiros e CREAs são **sempre pesquisados** automaticamente, 
    independentemente das palavras-chave adicionais inseridas.
    
    ### Sobre a verificação de assinatura:
    A verificação de assinatura é feita pelo nome do arquivo. 
    Arquivos que contêm "assinado" no nome (como "PRJ-FOT-IPER-04-05_assinado.pdf") 
    serão marcados como assinados na coluna "Arquivo assinado".
    
    ### Sobre o mapeamento de projetos:
    O código do projeto é extraído automaticamente del nome do arquivo (ex: "ECX" de "PRJ-ECX-IPER-02-07")
    e mapeado para a descrição correspondente. Em seguida, verifica-se se essa descrição está presente no conteúdo do PDF.
    """)

# Adicionar um rodapé
st.markdown("---")
st.markdown(
    """
    <style>
    .footer {
        text-align: center;
        color: gray;
        padding: 10px;
    }
    </style>
    <div class="footer">
    <p>Desenvolvido por Rodrigo Damasceno | © 2025 - Todos os direitos reservados</p>
    </div>
    """,
    unsafe_allow_html=True
)


